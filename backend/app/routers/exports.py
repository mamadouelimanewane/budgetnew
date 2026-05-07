from __future__ import annotations

import json
from datetime import datetime
from io import BytesIO
from pathlib import Path

from fastapi import APIRouter, Depends, HTTPException
from fastapi.responses import StreamingResponse
from pydantic import BaseModel, Field
from sqlalchemy import select
from sqlalchemy.orm import Session

from app.audit import append_audit_event
from app.deps import get_db
from app.models import BudgetAllocation, BudgetLine, Commitment, Payment, ReportRun, ReportSchedule, User
from app.routers.revenue import Revenue
from app.rbac import require_role


router = APIRouter(tags=["exports"])

REPORTS_DIR = Path(__file__).resolve().parents[2] / "models-store" / "reports"
REPORTS_DIR.mkdir(parents=True, exist_ok=True)


def _execution_rows(db: Session, *, budget_plan_id: int) -> list[dict]:
    # allocations
    allocs = db.scalars(select(BudgetAllocation).where(BudgetAllocation.budget_plan_id == budget_plan_id)).all()
    if not allocs:
        return []

    line_ids = {a.budget_line_id for a in allocs}
    lines = db.scalars(select(BudgetLine).where(BudgetLine.id.in_(line_ids))).all()
    line_by_id = {l.id: l for l in lines}

    # commitments + payments associated to plan (baseline linkage via commitment.budget_plan_id)
    commits = db.scalars(select(Commitment).where(Commitment.budget_plan_id == budget_plan_id)).all()
    commit_ids = [c.id for c in commits]
    payments = []
    if commit_ids:
        payments = db.scalars(select(Payment).where(Payment.commitment_id.in_(commit_ids))).all()

    paid_by_line_org: dict[tuple[int, str], int] = {}
    # Map payment->commitment to get budget_line/org_unit
    commit_by_id = {c.id: c for c in commits}
    for p in payments:
        c = commit_by_id.get(p.commitment_id)
        if not c:
            continue
        key = (c.budget_line_id, c.org_unit)
        paid_by_line_org[key] = paid_by_line_org.get(key, 0) + int(p.amount_xof)

    # revenues: baseline linkage via fiscal_year inferred from commitments
    fiscal_year = commits[0].fiscal_year if commits else None
    rev_by_org: dict[str, int] = {}
    if fiscal_year is not None:
        revs = db.scalars(select(Revenue).where(Revenue.fiscal_year == fiscal_year)).all()
        for r in revs:
            rev_by_org[r.org_unit] = rev_by_org.get(r.org_unit, 0) + int(r.amount_xof)

    rows = []
    for a in allocs:
        line = line_by_id.get(a.budget_line_id)
        paid = paid_by_line_org.get((a.budget_line_id, a.org_unit), 0)
        rows.append(
            {
                "org_unit": a.org_unit,
                "budget_line_code": line.code if line else str(a.budget_line_id),
                "budget_line_label": line.label if line else "",
                "allocated_xof": int(a.amount_xof),
                "paid_xof": int(paid),
                "available_xof": int(a.amount_xof) - int(paid),
                "consumption_rate": (float(paid) / float(a.amount_xof)) if a.amount_xof else 0.0,
                "revenue_xof_org": int(rev_by_org.get(a.org_unit, 0)),
            }
        )
    return rows


@router.get("/budget/execution.xlsx")
def export_budget_execution_xlsx(
    budget_plan_id: int,
    db: Session = Depends(get_db),
    user: User = Depends(require_role("admin", "analyst", "viewer")),
) -> StreamingResponse:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill

    rows = _execution_rows(db, budget_plan_id=budget_plan_id)
    if not rows:
        raise HTTPException(status_code=404, detail="No allocations for this budget_plan_id")

    wb = Workbook()
    ws = wb.active
    ws.title = "Execution"

    headers = [
        "OrgUnit",
        "LigneCode",
        "LigneLibellé",
        "Alloué (XOF)",
        "Payé (XOF)",
        "Disponible (XOF)",
        "Taux conso",
        "Recettes org (XOF)",
    ]
    ws.append(headers)
    header_fill = PatternFill("solid", fgColor="111827")
    header_font = Font(color="FFFFFF", bold=True)
    for c in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=c)
        cell.fill = header_fill
        cell.font = header_font

    for r in rows:
        ws.append(
            [
                r["org_unit"],
                r["budget_line_code"],
                r["budget_line_label"],
                r["allocated_xof"],
                r["paid_xof"],
                r["available_xof"],
                float(r["consumption_rate"]),
                r["revenue_xof_org"],
            ]
        )

    # simple formatting
    for col in (4, 5, 6, 8):
        for row in range(2, ws.max_row + 1):
            ws.cell(row=row, column=col).number_format = "#,##0"
    for row in range(2, ws.max_row + 1):
        ws.cell(row=row, column=7).number_format = "0.00%"

    out = BytesIO()
    wb.save(out)
    out.seek(0)

    append_audit_event(
        db=db,
        actor_user_id=user.id,
        action="export",
        entity="Report",
        entity_id="budget_execution_xlsx",
        details={"budget_plan_id": budget_plan_id},
    )

    headers_resp = {"Content-Disposition": f"attachment; filename=budget-execution-{budget_plan_id}.xlsx"}
    return StreamingResponse(out, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", headers=headers_resp)


@router.get("/budget/execution.pdf")
def export_budget_execution_pdf(
    budget_plan_id: int,
    db: Session = Depends(get_db),
    user: User = Depends(require_role("admin", "analyst", "viewer")),
) -> StreamingResponse:
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.pdfgen import canvas

    rows = _execution_rows(db, budget_plan_id=budget_plan_id)
    if not rows:
        raise HTTPException(status_code=404, detail="No allocations for this budget_plan_id")

    buf = BytesIO()
    c = canvas.Canvas(buf, pagesize=landscape(A4))
    width, height = landscape(A4)
    y = height - 30
    c.setFont("Helvetica-Bold", 14)
    c.drawString(30, y, f"État d'exécution budgétaire — BudgetPlan #{budget_plan_id}")
    y -= 18
    c.setFont("Helvetica", 9)
    c.drawString(30, y, f"Généré: {datetime.utcnow().isoformat(timespec='seconds')}Z")
    y -= 18

    c.setFont("Helvetica-Bold", 9)
    cols = ["Org", "Code", "Libellé", "Alloué", "Payé", "Dispo", "Taux"]
    x = [30, 120, 190, 520, 610, 700, 790]
    for i, name in enumerate(cols):
        c.drawString(x[i], y, name)
    y -= 12
    c.setFont("Helvetica", 8)

    for r in rows[:2500]:
        if y < 30:
            c.showPage()
            y = height - 30
            c.setFont("Helvetica", 8)
        c.drawString(x[0], y, str(r["org_unit"])[:12])
        c.drawString(x[1], y, str(r["budget_line_code"])[:10])
        c.drawString(x[2], y, str(r["budget_line_label"])[:45])
        c.drawRightString(x[3] + 60, y, f"{r['allocated_xof']:,}".replace(",", " "))
        c.drawRightString(x[4] + 60, y, f"{r['paid_xof']:,}".replace(",", " "))
        c.drawRightString(x[5] + 60, y, f"{r['available_xof']:,}".replace(",", " "))
        c.drawRightString(x[6] + 30, y, f"{r['consumption_rate']*100:.1f}%")
        y -= 10

    c.save()
    buf.seek(0)

    append_audit_event(
        db=db,
        actor_user_id=user.id,
        action="export",
        entity="Report",
        entity_id="budget_execution_pdf",
        details={"budget_plan_id": budget_plan_id},
    )

    headers_resp = {"Content-Disposition": f"attachment; filename=budget-execution-{budget_plan_id}.pdf"}
    return StreamingResponse(buf, media_type="application/pdf", headers=headers_resp)


class ScheduleIn(BaseModel):
    name: str = Field(min_length=1, max_length=200)
    report_type: str = Field(default="budget_execution", max_length=80)
    params: dict = Field(default_factory=dict)
    frequency: str = Field(default="manual", max_length=30)


class ScheduleOut(BaseModel):
    id: int
    name: str
    report_type: str
    params: dict
    frequency: str
    is_active: bool


@router.post("/schedules", response_model=ScheduleOut)
def create_schedule(
    payload: ScheduleIn,
    db: Session = Depends(get_db),
    user: User = Depends(require_role("admin", "analyst")),
) -> ScheduleOut:
    s = ReportSchedule(
        name=payload.name,
        report_type=payload.report_type,
        params_json=json.dumps(payload.params, sort_keys=True),
        frequency=payload.frequency,
        is_active=True,
        created_by_user_id=user.id,
    )
    db.add(s)
    db.commit()
    db.refresh(s)
    append_audit_event(db=db, actor_user_id=user.id, action="create", entity="ReportSchedule", entity_id=str(s.id), details=payload.model_dump())
    return ScheduleOut(id=s.id, name=s.name, report_type=s.report_type, params=payload.params, frequency=s.frequency, is_active=s.is_active)


@router.get("/schedules", response_model=list[ScheduleOut])
def list_schedules(
    db: Session = Depends(get_db),
    user: User = Depends(require_role("admin", "analyst")),
) -> list[ScheduleOut]:
    rows = db.scalars(select(ReportSchedule).order_by(ReportSchedule.id.desc()).limit(500)).all()
    out = []
    for r in rows:
        out.append(
            ScheduleOut(
                id=r.id,
                name=r.name,
                report_type=r.report_type,
                params=json.loads(r.params_json or "{}"),
                frequency=r.frequency,
                is_active=r.is_active,
            )
        )
    return out


class RunOut(BaseModel):
    id: int
    schedule_id: int
    status: str
    output_format: str
    file_path: str


@router.post("/schedules/{schedule_id}/run", response_model=RunOut)
def run_schedule(
    schedule_id: int,
    output_format: str = "xlsx",
    db: Session = Depends(get_db),
    user: User = Depends(require_role("admin", "analyst")),
) -> RunOut:
    s = db.get(ReportSchedule, schedule_id)
    if not s or not s.is_active:
        raise HTTPException(status_code=404, detail="Schedule not found")
    params = json.loads(s.params_json or "{}")
    report_type = s.report_type
    output_format = output_format.lower().strip()
    if output_format not in ("xlsx", "pdf"):
        raise HTTPException(status_code=400, detail="Unsupported format")

    if report_type != "budget_execution":
        raise HTTPException(status_code=400, detail="Unsupported report_type in v1")
    budget_plan_id = int(params.get("budget_plan_id", 0))
    if budget_plan_id <= 0:
        raise HTTPException(status_code=400, detail="Missing params.budget_plan_id")

    # generate file and write to disk (v1: build in-memory then persist)
    ts = datetime.utcnow().strftime("%Y%m%d-%H%M%S")
    file_path = REPORTS_DIR / f"schedule-{schedule_id}-run-{ts}.{output_format}"
    if output_format == "xlsx":
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill

        rows = _execution_rows(db, budget_plan_id=budget_plan_id)
        wb = Workbook()
        ws = wb.active
        ws.title = "Execution"
        headers = [
            "OrgUnit",
            "LigneCode",
            "LigneLibellé",
            "Alloué (XOF)",
            "Payé (XOF)",
            "Disponible (XOF)",
            "Taux conso",
            "Recettes org (XOF)",
        ]
        ws.append(headers)
        header_fill = PatternFill("solid", fgColor="111827")
        header_font = Font(color="FFFFFF", bold=True)
        for c in range(1, len(headers) + 1):
            cell = ws.cell(row=1, column=c)
            cell.fill = header_fill
            cell.font = header_font
        for r in rows:
            ws.append(
                [
                    r["org_unit"],
                    r["budget_line_code"],
                    r["budget_line_label"],
                    r["allocated_xof"],
                    r["paid_xof"],
                    r["available_xof"],
                    float(r["consumption_rate"]),
                    r["revenue_xof_org"],
                ]
            )
        for col in (4, 5, 6, 8):
            for row in range(2, ws.max_row + 1):
                ws.cell(row=row, column=col).number_format = "#,##0"
        for row in range(2, ws.max_row + 1):
            ws.cell(row=row, column=7).number_format = "0.00%"
        mem = BytesIO()
        wb.save(mem)
        content = mem.getvalue()
    else:
        from reportlab.lib.pagesizes import A4, landscape
        from reportlab.pdfgen import canvas

        rows = _execution_rows(db, budget_plan_id=budget_plan_id)
        mem = BytesIO()
        c = canvas.Canvas(mem, pagesize=landscape(A4))
        width, height = landscape(A4)
        y = height - 30
        c.setFont("Helvetica-Bold", 14)
        c.drawString(30, y, f"État d'exécution budgétaire — BudgetPlan #{budget_plan_id}")
        y -= 18
        c.setFont("Helvetica", 9)
        c.drawString(30, y, f"Généré: {datetime.utcnow().isoformat(timespec='seconds')}Z")
        y -= 18
        c.setFont("Helvetica-Bold", 9)
        cols = ["Org", "Code", "Libellé", "Alloué", "Payé", "Dispo", "Taux"]
        x = [30, 120, 190, 520, 610, 700, 790]
        for i, name in enumerate(cols):
            c.drawString(x[i], y, name)
        y -= 12
        c.setFont("Helvetica", 8)
        for r in rows[:2500]:
            if y < 30:
                c.showPage()
                y = height - 30
                c.setFont("Helvetica", 8)
            c.drawString(x[0], y, str(r["org_unit"])[:12])
            c.drawString(x[1], y, str(r["budget_line_code"])[:10])
            c.drawString(x[2], y, str(r["budget_line_label"])[:45])
            c.drawRightString(x[3] + 60, y, f"{r['allocated_xof']:,}".replace(",", " "))
            c.drawRightString(x[4] + 60, y, f"{r['paid_xof']:,}".replace(",", " "))
            c.drawRightString(x[5] + 60, y, f"{r['available_xof']:,}".replace(",", " "))
            c.drawRightString(x[6] + 30, y, f"{r['consumption_rate']*100:.1f}%")
            y -= 10
        c.save()
        content = mem.getvalue()
    file_path.write_bytes(content)

    run = ReportRun(schedule_id=schedule_id, status="completed", output_format=output_format, file_path=str(file_path))
    db.add(run)
    db.commit()
    db.refresh(run)
    append_audit_event(
        db=db,
        actor_user_id=user.id,
        action="run",
        entity="ReportSchedule",
        entity_id=str(schedule_id),
        details={"output_format": output_format, "file_path": str(file_path)},
    )
    return RunOut(id=run.id, schedule_id=run.schedule_id, status=run.status, output_format=run.output_format, file_path=run.file_path)


@router.get("/runs", response_model=list[RunOut])
def list_runs(
    schedule_id: int | None = None,
    db: Session = Depends(get_db),
    user: User = Depends(require_role("admin", "analyst")),
) -> list[RunOut]:
    stmt = select(ReportRun).order_by(ReportRun.id.desc()).limit(500)
    if schedule_id is not None:
        stmt = stmt.where(ReportRun.schedule_id == schedule_id)
    rows = db.scalars(stmt).all()
    return [RunOut(id=r.id, schedule_id=r.schedule_id, status=r.status, output_format=r.output_format, file_path=r.file_path) for r in rows]

